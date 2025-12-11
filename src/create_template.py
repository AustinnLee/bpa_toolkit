from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side


def create_financial_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly_Report"

    # 1. 设置标题 (模拟公司 Logo 区域)
    ws["A1"] = "BBA Global Sales Report - 2024"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(
        start_color="003366", end_color="003366", fill_type="solid"
    )

    # 2. 设置表头 (Row 3)
    headers = ["Region", "County", "Total Revenue", "Growth Rate"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # 3. 预设公式 (Row 10) - 假设数据只有 5 行
    ws["C10"] = "=SUM(C4:C9)"
    ws["C10"].font = Font(bold=True, color="FF0000")
    ws["B10"] = "Grand Total:"

    # 4. 保存
    root = Path(__file__).resolve().parent.parent
    path = root / "data" / "raw" / "financial_template.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(path)
    print(f"Template created at: {path}")


if __name__ == "__main__":
    create_financial_template()
