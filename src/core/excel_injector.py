from pathlib import Path
from typing import Union

import openpyxl
import pandas as pd


class ExcelInjector:
    def __init__(self, template_path: Union[str, Path]):
        self.template_path = Path(template_path)
        # 加载工作簿 (Workbook)
        print(f"Loading template: {self.template_path}")
        self.wb = openpyxl.load_workbook(self.template_path)

    def inject_dataframe(
        self, df: pd.DataFrame, sheet_name: str, start_row: int, start_col: int
    ):
        """
        核心方法：将 DataFrame 的数据逐个填入单元格。
        start_row: 从第几行开始填 (Excel行号，从1开始)
        start_col: 从第几列开始填 (Excel列号，A=1, B=2...)
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in template!")

        ws = self.wb[sheet_name]

        print(
            f"Injecting {len(df)} rows into sheet '{sheet_name}' starting at R{start_row}C{start_col}..."
        )

        # 将 DataFrame 转为 numpy array，遍历速度更快
        data_matrix = df.values.tolist()

        for i, row_data in enumerate(data_matrix):
            for j, value in enumerate(row_data):
                # 计算目标单元格位置
                target_row = start_row + i
                target_col = start_col + j

                # 获取单元格对象
                cell = ws.cell(row=target_row, column=target_col)

                # 关键：只修改 value，不动 style
                cell.value = value

        return self

    def save(self, output_path: Union[str, Path]):
        print(f"Saving report to: {output_path}")
        self.wb.save(output_path)
        print("Done.")


# --- 实战调用 ---
if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent

    # 1. 准备数据 (假设这是 DuckDB 算出来的结果)
    # 我们之前的 clean_real_sales.csv 里的数据
    # 为了演示简单，我们手动造一点数据，完全匹配模板的列数
    data = {
        "Region": ["North", "North", "South", "East", "West"],
        "County": ["Dallam", "Hartley", "Webb", "Shelby", "Pecos"],
        "Revenue": [15000, 23000, 4500, 12000, 8900],
        "Growth": [0.12, 0.05, -0.02, 0.08, 0.15],
    }
    df_result = pd.DataFrame(data)

    # 2. 实例化注入器
    template_file = root / "data" / "raw" / "financial_template.xlsx"
    injector = ExcelInjector(template_file)

    # 3. 注入数据
    # 模板表头在第3行，所以数据从第4行开始填
    # Region 在第1列 (A列)
    injector.inject_dataframe(
        df=df_result, sheet_name="Monthly_Report", start_row=4, start_col=1
    )

    # 4. 另存为新文件 (不要覆盖模板！)
    output_file = root / "data" / "processed" / "Final_Report_2024.xlsx"
    injector.save(output_file)
